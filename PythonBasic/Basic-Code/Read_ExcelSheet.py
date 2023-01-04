# Example 1- Read excel files in python using openpyxl

import openpyxl

# load workbook
wb = openpyxl.load_workbook("D:\\pythonProject\\API-Backend-Automation\\DATA\\Excel-Data.xlsx")
# this will return number of sheets
sheets = wb.sheetnames

# return sheet name which is active
print(wb.active.title)

# specify which sheet you would like to read
sh1 = wb['Sheet2']

# specify which cell to red
data = sh1['B2'].value

# option1
print(wb['Sheet2']['A2'].value)

# option2 which accept row and column
print(sh1.cell(3, 2).value)
print(sh1.cell(3, 1).value)
# here we are taking different sheet for example
